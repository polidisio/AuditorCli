import io
from datetime import datetime

import openpyxl
import pytest

from auditor.models import AuditSession, AuditTarget, Finding, Priority, Severity
from auditor.modules.report.excel import generate_excel


def _make_session() -> AuditSession:
    target = AuditTarget(domain="contoso.com", tenant_id="tenant-123")
    session = AuditSession(id="test01", target=target, started_at=datetime(2026, 6, 22, 10, 0))
    session.add_finding(Finding(
        id="M365-CA-001",
        title="Legacy Auth Not Blocked",
        component="Entra ID — Conditional Access",
        vector="SMTP AUTH bypasses MFA",
        mitre_id="T1078.004",
        severity=Severity.CRITICAL,
        priority=Priority.HIGH,
        description="Legacy authentication protocols enable MFA bypass.",
        remediation="Create CA policy blocking legacy auth clients.",
    ))
    session.add_finding(Finding(
        id="SPO-002",
        title="Default Sharing Link is Anonymous",
        component="SharePoint Online",
        vector="Anonymous link sharing by default",
        mitre_id="T1567.002",
        severity=Severity.MEDIUM,
        priority=Priority.MEDIUM,
        description="Default link type allows anonymous access.",
        evidence="Tenant setting: defaultSharingLinkType = anyone",
        remediation="Change to 'Specific people' in SharePoint Admin Center.",
    ))
    session.add_finding(Finding(
        id="TEAMS-004",
        title="Public Teams Visible to All",
        component="Teams — Team Visibility",
        vector="Any user can join public teams",
        mitre_id="T1087.004",
        severity=Severity.LOW,
        priority=Priority.LOW,
        description="3 teams are set to Public visibility.",
        remediation="Set sensitive teams to Private.",
    ))
    return session


def test_excel_generates_bytes():
    session = _make_session()
    data = generate_excel(session)
    assert isinstance(data, bytes)
    assert len(data) > 1000


def test_excel_three_sheets():
    session = _make_session()
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(session)))
    assert len(wb.sheetnames) == 3
    assert wb.sheetnames[0] == "Resumen Ejecutivo"
    assert wb.sheetnames[1] == "Hallazgos"
    assert wb.sheetnames[2] == "Matriz por Componente"


def test_excel_findings_row_count():
    session = _make_session()
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(session)))
    ws = wb["Hallazgos"]
    # header row + 3 findings = 4 rows with data
    data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in row)]
    assert len(data_rows) == 3


def test_excel_findings_sorted_by_priority():
    session = _make_session()
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(session)))
    ws = wb["Hallazgos"]
    # Column 6 = Prioridad
    priorities = [ws.cell(row=r, column=6).value for r in range(2, 5)]
    assert priorities == ["ALTA", "MEDIA", "BAJA"]


def test_excel_matrix_sheet_components():
    session = _make_session()
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(session)))
    ws = wb["Matriz por Componente"]
    components = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=1).value]
    assert len(components) == 3  # 3 unique components


def test_excel_summary_metadata():
    session = _make_session()
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(session)))
    ws = wb["Resumen Ejecutivo"]
    # Find the domain value anywhere in the sheet
    all_values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
    assert "contoso.com" in all_values


def test_excel_via_save_report(tmp_path):
    from auditor.modules.report.generator import save_report
    session = _make_session()
    path = save_report(session, tmp_path, fmt="xlsx")
    assert path.suffix == ".xlsx"
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert len(wb.sheetnames) == 3
