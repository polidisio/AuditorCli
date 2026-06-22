"""Excel (.xlsx) report generator using openpyxl."""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from auditor.models import AuditSession, Finding, Priority


# ── Palette ───────────────────────────────────────────────────────────────────
_FILL = {
    Priority.HIGH:   PatternFill("solid", fgColor="C0392B"),   # dark red
    Priority.MEDIUM: PatternFill("solid", fgColor="E67E22"),   # orange
    Priority.LOW:    PatternFill("solid", fgColor="2980B9"),   # blue
}
_FILL_LIGHT = {
    Priority.HIGH:   PatternFill("solid", fgColor="FADBD8"),
    Priority.MEDIUM: PatternFill("solid", fgColor="FDEBD0"),
    Priority.LOW:    PatternFill("solid", fgColor="D6EAF8"),
}
_FONT_WHITE = Font(color="FFFFFF", bold=True, size=10)
_FONT_DARK  = Font(color="1C1C1C", size=10)
_FONT_TITLE = Font(color="FFFFFF", bold=True, size=11)
_FONT_HDR   = Font(color="FFFFFF", bold=True, size=10)
_HEADER_FILL = PatternFill("solid", fgColor="1A252F")
_SECTION_FILL = PatternFill("solid", fgColor="2C3E50")
_THIN_BORDER = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)
_PRIORITY_ORDER = {Priority.HIGH: 0, Priority.MEDIUM: 1, Priority.LOW: 2}
_PRIORITY_LABEL = {Priority.HIGH: "ALTA", Priority.MEDIUM: "MEDIA", Priority.LOW: "BAJA"}


def _set_cell(ws, row: int, col: int, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    return cell


def _auto_col_width(ws, col: int, min_w: int = 10, max_w: int = 60) -> None:
    col_letter = get_column_letter(col)
    max_len = 0
    for cell in ws[col_letter]:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))


def _wrap_align(wrap: bool = True, v: str = "top", h: str = "left") -> Alignment:
    return Alignment(wrap_text=wrap, vertical=v, horizontal=h)


# ── Sheet 1: Resumen Ejecutivo ────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, session: AuditSession) -> None:
    ws = wb.active
    ws.title = "Resumen Ejecutivo"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36

    row = 1

    # Title banner
    ws.merge_cells(f"A{row}:B{row}")
    _set_cell(ws, row, 1, "INFORME DE AUDITORÍA DE SEGURIDAD",
              font=Font(color="FFFFFF", bold=True, size=14),
              fill=PatternFill("solid", fgColor="1A252F"),
              alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:B{row}")
    _set_cell(ws, row, 1, "AuditorCli — Reporte de Hallazgos",
              font=Font(color="BDC3C7", italic=True, size=10),
              fill=PatternFill("solid", fgColor="2C3E50"),
              alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 18
    row += 2

    # Metadata
    meta = [
        ("Target / Dominio", session.target.domain),
        ("Tenant ID", session.target.tenant_id or "—"),
        ("Session ID", session.id),
        ("Fecha de auditoría", session.started_at.strftime("%Y-%m-%d %H:%M UTC")),
        ("Generado", datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")),
        ("Herramienta", "AuditorCli v0.1"),
    ]
    for label, value in meta:
        _set_cell(ws, row, 1, label,
                  font=Font(bold=True, size=10, color="2C3E50"),
                  fill=PatternFill("solid", fgColor="ECF0F1"),
                  alignment=_wrap_align(False, "center"),
                  border=_THIN_BORDER)
        _set_cell(ws, row, 2, value,
                  font=_FONT_DARK,
                  alignment=_wrap_align(False, "center"),
                  border=_THIN_BORDER)
        row += 1

    row += 1

    # Counts section header
    ws.merge_cells(f"A{row}:B{row}")
    _set_cell(ws, row, 1, "RESUMEN DE HALLAZGOS",
              font=_FONT_TITLE,
              fill=_SECTION_FILL,
              alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 22
    row += 1

    counts = [
        ("Prioridad Alta",   session.high_count,   Priority.HIGH),
        ("Prioridad Media",  session.medium_count,  Priority.MEDIUM),
        ("Prioridad Baja",   session.low_count,     Priority.LOW),
        ("TOTAL",            len(session.findings), None),
    ]
    for label, count, priority in counts:
        fill = _FILL.get(priority, PatternFill("solid", fgColor="27AE60")) if priority else PatternFill("solid", fgColor="1A252F")
        font = _FONT_WHITE
        _set_cell(ws, row, 1, label, font=font, fill=fill,
                  alignment=_wrap_align(False, "center"), border=_THIN_BORDER)
        _set_cell(ws, row, 2, count,
                  font=Font(bold=True, size=12, color="FFFFFF"),
                  fill=fill,
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=_THIN_BORDER)
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # Legend
    ws.merge_cells(f"A{row}:B{row}")
    _set_cell(ws, row, 1, "LEYENDA DE PRIORIDADES",
              font=_FONT_TITLE,
              fill=_SECTION_FILL,
              alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 22
    row += 1

    legend = [
        ("ALTA",  "Requiere acción inmediata. Riesgo crítico o alto impacto.",  Priority.HIGH),
        ("MEDIA", "Requiere acción planificada. Riesgo significativo.",          Priority.MEDIUM),
        ("BAJA",  "Mejora recomendada. Impacto limitado o difícil explotación.", Priority.LOW),
    ]
    for label, desc, priority in legend:
        _set_cell(ws, row, 1, label,
                  font=_FONT_WHITE,
                  fill=_FILL[priority],
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=_THIN_BORDER)
        _set_cell(ws, row, 2, desc,
                  font=_FONT_DARK,
                  alignment=_wrap_align(True, "center"),
                  border=_THIN_BORDER)
        ws.row_dimensions[row].height = 18
        row += 1


# ── Sheet 2: Hallazgos ────────────────────────────────────────────────────────

_FINDINGS_COLS = [
    ("ID",             12),
    ("Componente",     22),
    ("Título",         36),
    ("MITRE ATT&CK",   14),
    ("Severidad",      12),
    ("Prioridad",      11),
    ("Vector",         34),
    ("Descripción",    46),
    ("Evidencia",      34),
    ("Remediación",    46),
]


def _build_findings_sheet(wb: Workbook, session: AuditSession) -> None:
    ws = wb.create_sheet("Hallazgos")
    ws.sheet_view.showGridLines = False

    # Column widths
    for col_idx, (_, width) in enumerate(_FINDINGS_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Header row
    for col_idx, (header, _) in enumerate(_FINDINGS_COLS, start=1):
        _set_cell(ws, 1, col_idx, header,
                  font=_FONT_HDR,
                  fill=_HEADER_FILL,
                  alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                  border=_THIN_BORDER)
    ws.row_dimensions[1].height = 22

    # Freeze header
    ws.freeze_panes = "A2"

    # AutoFilter
    last_col = get_column_letter(len(_FINDINGS_COLS))
    ws.auto_filter.ref = f"A1:{last_col}1"

    # Data rows — sorted by priority
    sorted_findings = sorted(session.findings, key=lambda f: _PRIORITY_ORDER.get(f.priority, 99))

    for row_idx, finding in enumerate(sorted_findings, start=2):
        priority = finding.priority
        row_fill = _FILL_LIGHT[priority]
        label = _PRIORITY_LABEL.get(priority, priority.value.upper())

        values = [
            finding.id,
            finding.component,
            finding.title,
            finding.mitre_id or "—",
            finding.severity.value.upper(),
            label,
            finding.vector,
            finding.description,
            finding.evidence or "—",
            finding.remediation,
        ]

        for col_idx, value in enumerate(values, start=1):
            is_text_col = col_idx in (7, 8, 9, 10)  # vector, desc, evidence, remediation
            _set_cell(
                ws, row_idx, col_idx, value,
                font=_FONT_DARK,
                fill=row_fill,
                alignment=_wrap_align(wrap=is_text_col, v="top"),
                border=_THIN_BORDER,
            )

        # Priority cell gets solid color
        _set_cell(ws, row_idx, 6, label,
                  font=_FONT_WHITE,
                  fill=_FILL[priority],
                  alignment=Alignment(horizontal="center", vertical="top"),
                  border=_THIN_BORDER)

        ws.row_dimensions[row_idx].height = 60 if any([
            len(str(finding.description)) > 80,
            len(str(finding.remediation)) > 80,
        ]) else 40


# ── Sheet 3: Matriz por Componente ────────────────────────────────────────────

def _build_matrix_sheet(wb: Workbook, session: AuditSession) -> None:
    ws = wb.create_sheet("Matriz por Componente")
    ws.sheet_view.showGridLines = False

    headers = ["Componente", "Alta", "Media", "Baja", "Total"]
    col_widths = [40, 10, 10, 10, 10]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        _set_cell(ws, 1, col_idx, header,
                  font=_FONT_HDR,
                  fill=_HEADER_FILL,
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=_THIN_BORDER)
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Aggregate by component
    counts: dict[str, dict[Priority, int]] = defaultdict(lambda: {p: 0 for p in Priority})
    for f in session.findings:
        counts[f.component][f.priority] += 1

    sorted_components = sorted(
        counts.items(),
        key=lambda x: (-(x[1][Priority.HIGH]), -(x[1][Priority.MEDIUM]), -(x[1][Priority.LOW]))
    )

    for row_idx, (component, prio_counts) in enumerate(sorted_components, start=2):
        high   = prio_counts[Priority.HIGH]
        medium = prio_counts[Priority.MEDIUM]
        low    = prio_counts[Priority.LOW]
        total  = high + medium + low

        # Row background based on dominant priority
        if high > 0:
            row_fill = _FILL_LIGHT[Priority.HIGH]
        elif medium > 0:
            row_fill = _FILL_LIGHT[Priority.MEDIUM]
        else:
            row_fill = _FILL_LIGHT[Priority.LOW]

        _set_cell(ws, row_idx, 1, component, font=_FONT_DARK, fill=row_fill,
                  alignment=_wrap_align(True, "center"), border=_THIN_BORDER)

        count_data = [
            (2, high,   Priority.HIGH),
            (3, medium, Priority.MEDIUM),
            (4, low,    Priority.LOW),
        ]
        for col, count, priority in count_data:
            cell_fill = _FILL[priority] if count > 0 else PatternFill("solid", fgColor="ECF0F1")
            cell_font = _FONT_WHITE if count > 0 else Font(color="95A5A6", size=10)
            _set_cell(ws, row_idx, col, count,
                      font=cell_font, fill=cell_fill,
                      alignment=Alignment(horizontal="center", vertical="center"),
                      border=_THIN_BORDER)

        _set_cell(ws, row_idx, 5, total,
                  font=Font(bold=True, size=10, color="1A252F"),
                  fill=PatternFill("solid", fgColor="D5D8DC"),
                  alignment=Alignment(horizontal="center", vertical="center"),
                  border=_THIN_BORDER)

        ws.row_dimensions[row_idx].height = 22


# ── Public API ────────────────────────────────────────────────────────────────

def generate_excel(session: AuditSession) -> bytes:
    """Return xlsx file as bytes (for streaming or testing)."""
    wb = Workbook()
    _build_summary_sheet(wb, session)
    _build_findings_sheet(wb, session)
    _build_matrix_sheet(wb, session)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_excel(session: AuditSession, output_dir: Path) -> Path:
    """Write .xlsx to output_dir and return the path."""
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    domain_slug = session.target.domain.replace(".", "_")
    path = output_dir / f"audit_{domain_slug}_{ts}.xlsx"
    path.write_bytes(generate_excel(session))
    return path
